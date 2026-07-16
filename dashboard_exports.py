"""Geração de dashboards executivos em XLSX e PNG."""

from __future__ import annotations

import io
from dataclasses import asdict
from typing import Any, Sequence

import pandas as pd

from analytics_engine import FutureProjection, build_future_projection, generate_local_intelligent_report
from dominio_importers import (
    DominioSimulationReport,
    MonthlyReport,
    PGDASReport,
    cnpjs_are_compatible,
)
from simples_lc214 import SimplesLC214Simulation
from nascel_consulting import (
    NASCEL_COLORS,
    NASCEL_NAME,
    NASCEL_TAGLINE,
    build_nascel_diagnostic,
    decision_matrix_frame,
    legal_timeline_frame,
    official_sources_frame,
)


COLORS = {
    "navy": NASCEL_COLORS["navy"],
    "blue": NASCEL_COLORS["navy"],
    "green": NASCEL_COLORS["green"],
    "yellow": NASCEL_COLORS["gold"],
    "red": NASCEL_COLORS["red"],
    "slate": NASCEL_COLORS["slate"],
    "light": NASCEL_COLORS["cream"],
}


def scenario_table(
    report: DominioSimulationReport,
    lc214_simulation: SimplesLC214Simulation | None = None,
) -> pd.DataFrame:
    revenue = report.base_saidas
    current_total = report.tributos_atuais["Total"]
    effective_rate = current_total / revenue if revenue else 0.0
    # A comparação isola o efeito da reforma: com a mesma base, Anexo e
    # segregações, 2027 por dentro preserva a carga efetiva de 2026. A tabela
    # legal é usada para a partilha interna, não para inventar uma diferença
    # causada por RBT12 ou premissas incompatíveis com o arquivo importado.
    inside_total = current_total
    inside_effective_rate = effective_rate
    inside_label = "Simples Por Dentro 2027 — mesma carga de 2026"
    credit_2027 = report.base_entradas_credito * (
        report.aliquota_credito_cbs_2027 + report.aliquota_credito_ibs_2027
    )
    credit_2033 = report.base_entradas_credito * (
        report.aliquota_credito_cbs_2033 + report.aliquota_credito_ibs_2033
    )
    rows = [
        {
            "Cenário": "Simples 2026 — atual",
            "Carga Tributária": current_total,
            "Carga Efetiva": effective_rate,
            "Crédito Estimado das Compras": 0.0,
            "Variação vs. Atual": 0.0,
            "Leitura": "Referência atual",
        },
        {
            "Cenário": inside_label,
            "Carga Tributária": inside_total,
            "Carga Efetiva": inside_effective_rate,
            "Crédito Estimado das Compras": 0.0,
            "Variação vs. Atual": inside_total - current_total,
            "Leitura": "Sem diferença de tributação vs. 2026 na mesma base; muda somente a partilha do DAS",
        },
        {
            "Cenário": "Híbrido 2027 — Domínio",
            "Carga Tributária": report.fase_2027["total"],
            "Carga Efetiva": report.fase_2027["total"] / revenue if revenue else 0.0,
            "Crédito Estimado das Compras": credit_2027,
            "Variação vs. Atual": report.fase_2027["diferenca"],
            "Leitura": "Crédito das compras sujeito a documento idôneo e demais requisitos",
        },
        {
            "Cenário": "Híbrido 2033 — Domínio",
            "Carga Tributária": report.fase_2033["total"],
            "Carga Efetiva": report.fase_2033["total"] / revenue if revenue else 0.0,
            "Crédito Estimado das Compras": credit_2033,
            "Variação vs. Atual": report.fase_2033["diferenca"],
            "Leitura": "Menor carga estimada, sujeita à confirmação das alíquotas futuras",
        },
    ]
    return pd.DataFrame(rows)


def attention_points(
    report: DominioSimulationReport,
    monthly: MonthlyReport,
    pgdas: PGDASReport | None = None,
) -> list[dict[str, str]]:
    period_rows = monthly.movimentos[
        monthly.movimentos["Competência"].dt.to_period("M") == report.periodo.to_period("M")
    ]
    monthly_inputs = float(period_rows["Entradas"].sum()) if not period_rows.empty else 0.0
    reconciliation = monthly_inputs - report.base_entradas_credito
    no_input_activity = monthly_inputs == 0 and report.base_entradas_credito == 0
    points: list[dict[str, str]] = []
    if pgdas is not None and not cnpjs_are_compatible(report.cnpj, pgdas):
        points.append(
            {
                "Prioridade": "CRÍTICO",
                "Tema": "CNPJ divergente no PGDAS",
                "O que significa": "Os arquivos não representam a mesma empresa e não podem sustentar uma conclusão conjunta.",
                "Dados considerados": "CNPJ da Simulação da Reforma e CNPJ do extrato PGDAS-D.",
                "Constatação": f"Domínio: {report.cnpj} | PGDAS: {pgdas.cnpj_estabelecimento or pgdas.cnpj_basico}",
                "Ação": "Não consolidar os arquivos. Importar o PGDAS da mesma empresa antes da decisão.",
            }
        )
    if abs(reconciliation) >= 0.01:
        points.append(
            {
                "Prioridade": "ATENÇÃO",
                "Tema": "Conciliação das entradas",
                "O que significa": "A base usada para crédito não coincide com o total de entradas do mesmo mês.",
                "Dados considerados": "Entradas do Demonstrativo Mensal menos a base de crédito da Simulação da Reforma.",
                "Constatação": f"Demonstrativo mensal menos base de crédito: R$ {reconciliation:,.2f}",
                "Ação": "Identificar acumuladores excluídos e confirmar quais aquisições geram crédito.",
            }
        )
    input_point = (
        {
            "Prioridade": "DECISÃO",
            "Tema": "Empresa sem entradas na competência",
            "O que significa": "A ausência da seção de entradas é compatível com uma prestadora sem compras no período.",
            "Dados considerados": "Demonstrativo Mensal e base de créditos pelas entradas da Simulação da Reforma.",
            "Constatação": "Entradas contábeis e base de crédito iguais a R$ 0,00.",
            "Ação": "Confirmar a ausência de compras e manter os créditos de IBS/CBS zerados.",
        }
        if no_input_activity
        else {
            "Prioridade": "CRÍTICO",
            "Tema": "Elegibilidade dos créditos",
            "O que significa": "Nem toda entrada contábil necessariamente gera crédito aproveitável de IBS/CBS.",
            "Dados considerados": "Base de entradas indicada pelo Domínio, antes da validação individual de documentos e operações.",
            "Constatação": f"A simulação usa R$ {report.base_entradas_credito:,.2f} como base de entradas.",
            "Ação": "Revisar brindes, uso/consumo, documentos inidôneos, pagamentos e operações com tratamento específico.",
        }
    )
    points.extend(
        [
            {
                "Prioridade": "ATENÇÃO",
                "Tema": "Dependência B2B estimada",
                "O que significa": "Quanto maior a venda para empresas do regime regular, maior pode ser a importância do crédito transferido ao cliente.",
                "Dados considerados": "Saídas totais menos o acumulador classificado como não contribuinte; é uma aproximação, não cadastro cliente a cliente.",
                "Constatação": f"{report.percentual_operacoes_creditaveis:.2%} das saídas não estão no acumulador 'não contribuinte'.",
                "Ação": "Validar com relatório por CNPJ/CPF; o acumulador é apenas uma proxy de crédito comercial.",
            },
            input_point,
            {
                "Prioridade": "ATENÇÃO",
                "Tema": "Alíquotas futuras",
                "O que significa": "O resultado de longo prazo muda quando as alíquotas de referência ou tratamentos da operação forem atualizados.",
                "Dados considerados": "Alíquotas CBS e IBS cadastradas no relatório Domínio para o cenário 2033.",
                "Constatação": (
                    f"Domínio: CBS {report.aliquota_cbs_2033:.2%} e IBS {report.aliquota_ibs_2033:.2%} em 2033."
                ),
                "Ação": "Atualizar a simulação quando as alíquotas de referência e do destino forem publicadas.",
            },
            {
                "Prioridade": "DECISÃO",
                "Tema": "Opção 2027",
                "O que significa": "Compara o custo tributário estimado do regime regular de IBS/CBS com a permanência integral no Simples.",
                "Dados considerados": "DAS residual, CBS, IBS e total calculados pelo relatório Domínio para 2027.",
                "Constatação": f"Híbrido varia R$ {report.fase_2027['diferenca']:,.2f} ({report.fase_2027['diferenca_percentual']:.2%}).",
                "Ação": "Comparar o pequeno impacto de carga com retenção de clientes B2B e custo operacional.",
            },
            {
                "Prioridade": "DECISÃO",
                "Tema": "Cenário estrutural 2033",
                "O que significa": "Mostra a direção econômica esperada quando a transição estiver concluída.",
                "Dados considerados": "DAS residual, CBS, IBS, créditos e alíquotas do cenário 2033 informado pelo Domínio.",
                "Constatação": f"Híbrido varia R$ {report.fase_2033['diferenca']:,.2f} ({report.fase_2033['diferenca_percentual']:.2%}).",
                "Ação": "Planejar cadastro de itens, fornecedores, destinos, documentos e conciliação de créditos.",
            },
        ]
    )
    return points


def dashboard_explanations(
    report: DominioSimulationReport,
    monthly: MonthlyReport,
    pgdas: PGDASReport | None,
    projection: FutureProjection,
    lc214_simulation: SimplesLC214Simulation | None = None,
) -> pd.DataFrame:
    """Explica indicadores, cálculos e limitações em linguagem gerencial."""
    compatible_pgdas = bool(pgdas and cnpjs_are_compatible(report.cnpj, pgdas))
    period_rows = monthly.movimentos[
        monthly.movimentos["Competência"].dt.to_period("M") == report.periodo.to_period("M")
    ]
    monthly_inputs = float(period_rows["Entradas"].sum()) if not period_rows.empty else 0.0
    reconciliation = monthly_inputs - report.base_entradas_credito
    no_input_activity = monthly_inputs == 0 and report.base_entradas_credito == 0
    rows = [
        {
            "Indicador": "Saídas analisadas",
            "O que mostra": "Receita de vendas/serviços usada como base da competência selecionada.",
            "Como foi obtido": "Base de débitos pelas saídas no relatório Simulação da Reforma.",
            "Fonte ou premissa": "Domínio · competência mais recente importada.",
            "Como interpretar": "É o denominador das cargas efetivas; não representa caixa ou lucro.",
        },
        {
            "Indicador": "Base de entradas para crédito",
            "O que mostra": "Ausência de aquisições e de créditos na competência."
            if no_input_activity else "Montante de aquisições considerado potencialmente creditável.",
            "Como foi obtido": "Base de créditos pelas entradas indicada na Simulação da Reforma.",
            "Fonte ou premissa": "Domínio; depende de documento, operação e elegibilidade fiscal.",
            "Como interpretar": "Valor zero é válido para empresa sem entradas; confirme a ausência ao atualizar a competência."
            if no_input_activity else "Não deve ser tratada como crédito definitivo antes da conciliação fiscal.",
        },
        {
            "Indicador": "Compras potencialmente creditáveis",
            "O que mostra": "Não aplicável: não há compras na competência."
            if no_input_activity else "Parcela das compras atuais considerada na base de crédito de IBS/CBS.",
            "Como foi obtido": "Definido como zero quando entradas e base de crédito são ambas nulas."
            if no_input_activity else "Base de crédito das entradas ÷ entradas contábeis da mesma competência.",
            "Fonte ou premissa": "Simulação da Reforma e Demonstrativo Mensal do Domínio.",
            "Como interpretar": "Sem compras, não há crédito de entradas para reduzir CBS/IBS no Híbrido."
            if no_input_activity else "Percentual alto aumenta o crédito potencial no Híbrido, mas exige validação por item e documento.",
        },
        {
            "Indicador": "Atual / Por Dentro",
            "O que mostra": "2026 e 2027 Por Dentro com a mesma carga efetiva quando a base, o Anexo e as segregações não mudam.",
            "Como foi obtido": "Alíquota efetiva importada de 2026 aplicada à mesma base; em 2027 muda apenas a repartição interna do DAS.",
            "Fonte ou premissa": f"Tributos atuais do {'PGDAS-D/Domínio conciliados' if compatible_pgdas else 'Domínio'}; tabela legal usada para a partilha.",
            "Como interpretar": "Diferença zero na mesma base. Uma variação só deve surgir por mudança de receita, faixa, Anexo ou segregação.",
        },
        {
            "Indicador": "Híbrido 2027 / Por Fora",
            "O que mostra": "Custo de manter o Simples para os demais tributos e apurar CBS/IBS no regime regular.",
            "Como foi obtido": "DAS residual + CBS líquida + IBS líquido calculados pelo Domínio.",
            "Fonte ou premissa": "Cenário 2027 do relatório Simulação da Reforma e base de entradas para créditos.",
            "Como interpretar": "Compare a diferença de carga com o benefício comercial dos créditos e o custo de controle.",
        },
        {
            "Indicador": "Híbrido 2033",
            "O que mostra": "Cenário estrutural após a transição, sujeito às alíquotas e regras futuras.",
            "Como foi obtido": "DAS residual + CBS + IBS líquidos projetados pelo Domínio para 2033.",
            "Fonte ou premissa": "Alíquotas 2033 existentes no arquivo importado.",
            "Como interpretar": "É uma direção de planejamento, não uma apuração definitiva hoje.",
        },
        {
            "Indicador": "Projeção anual 2027 e 2033",
            "O que mostra": "Totais do ano-calendário de 2027 e cenário estrutural de 2033 com a mesma base anual.",
            "Como foi obtido": f"Média dos últimos {projection.meses_media} meses, crescimento anual de {projection.crescimento_anual:.2%} e alíquota Por Dentro de {projection.aliquota_por_dentro:.2%}.",
            "Fonte ou premissa": f"Demonstrativo Mensal; crescimento {'automático' if projection.modo_crescimento == 'average' else 'informado pelo usuário'}.",
            "Como interpretar": "2033 mantém receita e compras de 2027 para que a diferença mostre o efeito tributário, não crescimento comercial de longo prazo.",
        },
        {
            "Indicador": "Créditos na projeção",
            "O que mostra": "Crédito estimado de CBS/IBS que a empresa poderá aproveitar nas compras.",
            "Como foi obtido": "Compras projetadas × proporção atual da base creditável × alíquotas de crédito das entradas importadas do Domínio.",
            "Fonte ou premissa": "Linha Créditos pelas Entradas da Simulação da Reforma.",
            "Como interpretar": "No DAS Normal o crédito é zero; no Híbrido ele reduz CBS e IBS separadamente, sujeito à validação fiscal.",
        },
        {
            "Indicador": "Sensibilidade das compras",
            "O que mostra": "Se o Híbrido 2033 continua vantajoso quando a base de compras creditáveis muda.",
            "Como foi obtido": "Mantém a receita e o DAS Normal da projeção anual e recalcula CBS/IBS para bases creditáveis de 50%, 70%, atual e 100% da receita.",
            "Fonte ou premissa": "Alíquotas de débitos e créditos de 2033 importadas do Domínio.",
            "Como interpretar": "Diferença negativa favorece o Híbrido; positiva favorece o DAS Normal. Compare a base atual com o ponto de equilíbrio.",
        },
        {
            "Indicador": "Conciliação das entradas",
            "O que mostra": "Diferença entre o movimento contábil de entradas e a base usada para crédito.",
            "Como foi obtido": "Entradas do Demonstrativo Mensal - base de crédito da Simulação da Reforma.",
            "Fonte ou premissa": f"Diferença da competência: R$ {reconciliation:,.2f}.",
            "Como interpretar": "Conciliação válida em zero; confirme que não houve compras no período."
            if no_input_activity else "Diferença relevante deve ser explicada antes de decidir pelo regime por fora.",
        },
    ]
    if lc214_simulation is not None:
        rows.append(
            {
                "Indicador": "Nova tabela LC 214/2025",
                "O que mostra": f"Simples 2027/2028 no Anexo {lc214_simulation.annex}, {lc214_simulation.bracket}ª faixa.",
                "Como foi obtido": "(RBT12 × alíquota nominal - parcela a deduzir) ÷ RBT12, com repartição por tributo.",
                "Fonte ou premissa": f"RBT12 de R$ {lc214_simulation.rbt12:,.2f}; taxa teórica da tabela {lc214_simulation.effective_rate:.2%}.",
                "Como interpretar": "A taxa teórica serve para conferir faixa e partilha. A comparação 2026/2027 preserva a alíquota efetiva importada. No Anexo II, o IPI é mantido até norma específica em sentido diverso.",
            }
        )
    return pd.DataFrame(rows)


def create_dashboard_images(
    report: DominioSimulationReport,
    monthly: MonthlyReport,
    pgdas: PGDASReport | None = None,
    projection: FutureProjection | None = None,
    lc214_simulation: SimplesLC214Simulation | None = None,
) -> dict[str, bytes]:
    import os
    import tempfile
    from pathlib import Path

    cache_dir = Path(tempfile.gettempdir()) / "portal_ibs_cbs_matplotlib"
    cache_dir.mkdir(parents=True, exist_ok=True)
    os.environ.setdefault("MPLCONFIGDIR", str(cache_dir))
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    scenarios = scenario_table(report, lc214_simulation)
    projection = projection or build_future_projection(
        [report], monthly, lc214_simulation=lc214_simulation
    )
    images: dict[str, bytes] = {}

    fig, ax1 = plt.subplots(figsize=(12, 6.5), facecolor="white")
    names = ["2026 atual", "2027 Por Dentro\n(mesma carga)", "Híbrido\n2027", "Híbrido\n2033"]
    x = range(len(names))
    bars = ax1.bar(
        x,
        scenarios["Carga Tributária"],
        color=[COLORS["slate"], COLORS["blue"], COLORS["green"], COLORS["yellow"]],
        width=0.62,
    )
    ax1.set_xticks(list(x), names)
    ax1.set_ylabel("Carga tributária (R$)", color=COLORS["navy"])
    ax1.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"R$ {value:,.0f}"))
    ax1.grid(axis="y", alpha=0.18)
    ax1.set_axisbelow(True)
    for bar, value in zip(bars, scenarios["Carga Tributária"]):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 55, f"R$ {value:,.2f}", ha="center", fontsize=9)
    ax2 = ax1.twinx()
    ax2.plot(
        list(x), scenarios["Crédito Estimado das Compras"], color=COLORS["red"], marker="o", linewidth=2.5,
        label="Crédito estimado das compras",
    )
    ax2.set_ylabel("Crédito estimado das compras (R$)", color=COLORS["red"])
    ax2.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"R$ {value:,.0f}"))
    fig.suptitle("Comparativo dos cenários tributários", fontsize=18, fontweight="bold", color=COLORS["navy"])
    ax1.set_title(f"{report.empresa} · competência {report.periodo:%m/%Y}", color=COLORS["slate"])
    fig.tight_layout()
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    plt.close(fig)
    images["dashboard_cenarios.png"] = buffer.getvalue()

    movements = monthly.movimentos.copy()
    fig, ax = plt.subplots(figsize=(12, 6.5), facecolor="white")
    ax.plot(movements["Competência"], movements["Saídas"], marker="o", color=COLORS["blue"], label="Saídas")
    ax.plot(movements["Competência"], movements["Entradas"], marker="o", color=COLORS["green"], label="Entradas")
    ax.fill_between(movements["Competência"], movements["Saídas"], alpha=0.08, color=COLORS["blue"])
    ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"R$ {value/1000:.0f} mil"))
    ax.grid(alpha=0.2)
    ax.legend(frameon=False, ncol=2)
    ax.set_title("Evolução mensal de entradas e saídas", fontsize=18, fontweight="bold", color=COLORS["navy"])
    ax.set_xlabel("Competência")
    ax.set_ylabel("Movimentação")
    fig.autofmt_xdate()
    fig.tight_layout()
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    plt.close(fig)
    images["dashboard_evolucao_mensal.png"] = buffer.getvalue()

    points = attention_points(report, monthly, pgdas)
    fig, ax = plt.subplots(figsize=(13, 7.5), facecolor="white")
    ax.axis("off")
    ax.text(0.02, 0.95, "Pontos de atenção para a decisão", fontsize=21, fontweight="bold", color=COLORS["navy"])
    ax.text(0.02, 0.90, report.empresa, fontsize=12, color=COLORS["slate"])
    y = 0.82
    priority_colors = {"CRÍTICO": COLORS["red"], "ATENÇÃO": "#D68910", "DECISÃO": COLORS["blue"]}
    for point in points[:6]:
        color = priority_colors.get(point["Prioridade"], COLORS["slate"])
        ax.text(0.02, y, point["Prioridade"], fontsize=9, fontweight="bold", color="white", bbox=dict(boxstyle="round,pad=0.35", facecolor=color, edgecolor="none"))
        ax.text(0.15, y + 0.006, point["Tema"], fontsize=12, fontweight="bold", color=COLORS["navy"])
        ax.text(0.15, y - 0.035, point["Constatação"], fontsize=9.5, color=COLORS["slate"])
        ax.text(0.15, y - 0.068, "Ação: " + point["Ação"], fontsize=9.5, color="#263238", wrap=True)
        y -= 0.13
    fig.tight_layout()
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    plt.close(fig)
    images["dashboard_pontos_atencao.png"] = buffer.getvalue()

    annual = projection.resumo_anual
    fig, ax = plt.subplots(figsize=(12, 6.5), facecolor="white")
    x = list(range(len(annual)))
    normal_positions = [value - 0.2 for value in x]
    hybrid_positions = [value + 0.2 for value in x]
    normal_bars = ax.bar(
        normal_positions,
        annual["DAS Normal · Valor"],
        width=0.38,
        color=COLORS["blue"],
        label="DAS Normal",
    )
    hybrid_bars = ax.bar(
        hybrid_positions,
        annual["Híbrido · Total a Pagar"],
        width=0.38,
        color=COLORS["green"],
        label="Híbrido",
    )
    for bars, rates in (
        (normal_bars, annual["DAS Normal · Alíquota Efetiva"]),
        (hybrid_bars, annual["Híbrido · Alíquota Efetiva"]),
    ):
        for bar, rate in zip(bars, rates):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{rate:.2%}",
                ha="center",
                va="bottom",
                fontsize=10,
            )
    ax.set_xticks(x, annual["Período"])
    ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"R$ {value:,.0f}"))
    ax.grid(axis="y", alpha=0.2)
    ax.legend(frameon=False, ncol=2)
    ax.set_title(
        "Projeção anual simplificada — total a pagar e alíquota efetiva",
        fontsize=18,
        fontweight="bold",
        color=COLORS["navy"],
    )
    ax.set_xlabel("2027 completo e 2033 com a mesma base de 2027")
    ax.set_ylabel("Total anual estimado")
    fig.tight_layout()
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png", dpi=180, bbox_inches="tight")
    plt.close(fig)
    images["dashboard_projecao_futura.png"] = buffer.getvalue()
    return images


def _build_excel_dashboard_openpyxl(
    report: DominioSimulationReport,
    monthly: MonthlyReport,
    pgdas: PGDASReport | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.drawing.image import Image
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    ws_dashboard = wb.create_sheet("Dashboard")
    ws_scenarios = wb.create_sheet("Cenarios")
    ws_sensitivity = wb.create_sheet("Sensibilidade")
    ws_monthly = wb.create_sheet("Evolucao_Mensal")
    ws_composition = wb.create_sheet("Composicao")
    ws_attention = wb.create_sheet("Pontos_Atencao")
    ws_raw = wb.create_sheet("Dados_Dominio")
    ws_assumptions = wb.create_sheet("Premissas")
    if pgdas is not None:
        ws_pgdas = wb.create_sheet("PGDAS_Exemplo")

    navy_fill = PatternFill("solid", fgColor="123B5D")
    blue_fill = PatternFill("solid", fgColor="DCEAF7")
    green_fill = PatternFill("solid", fgColor="DFF3EC")
    yellow_fill = PatternFill("solid", fgColor="FFF4CC")
    red_fill = PatternFill("solid", fgColor="FDE2E0")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=22, bold=True, color="123B5D")
    subtitle_font = Font(size=11, color="526579")
    thin = Side(style="thin", color="D7E0E8")
    money_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
    percent_format = "0.00%"

    def style_header(sheet: Any, row: int, start: int, end: int) -> None:
        for column in range(start, end + 1):
            cell = sheet.cell(row=row, column=column)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)

    def autofit(sheet: Any, maximum: int = 45) -> None:
        for column_cells in sheet.columns:
            width = min(max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells) + 2, maximum)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(width, 11)

    scenarios = scenario_table(report)
    points = attention_points(report, monthly, pgdas)
    images = create_dashboard_images(report, monthly, pgdas)
    period_row = monthly.movimentos[
        monthly.movimentos["Competência"].dt.to_period("M") == report.periodo.to_period("M")
    ]
    monthly_inputs = float(period_row["Entradas"].sum()) if not period_row.empty else 0.0
    reconciliation = monthly_inputs - report.base_entradas_credito

    ws_dashboard.sheet_view.showGridLines = False
    ws_dashboard.merge_cells("A1:L2")
    ws_dashboard["A1"] = "Dashboard Executivo — Reforma Tributária IBS/CBS"
    ws_dashboard["A1"].font = title_font
    ws_dashboard["A1"].alignment = Alignment(vertical="center")
    ws_dashboard.merge_cells("A3:L3")
    ws_dashboard["A3"] = f"{report.empresa} | CNPJ {report.cnpj} | Competência {report.periodo:%m/%Y}"
    ws_dashboard["A3"].font = subtitle_font
    kpis = [
        ("A5", "C5", "Saídas analisadas", report.base_saidas, money_format, blue_fill),
        ("D5", "F5", "Base de entradas", report.base_entradas_credito, money_format, green_fill),
        ("G5", "I5", "Operações potencialmente creditáveis", report.percentual_operacoes_creditaveis, percent_format, yellow_fill),
        ("J5", "L5", "Diferença 2027", report.fase_2027["diferenca_percentual"], percent_format, red_fill if report.fase_2027["diferenca"] > 0 else green_fill),
    ]
    for start, end, label, value, number_format, fill in kpis:
        ws_dashboard.merge_cells(f"{start}:{end}")
        cell = ws_dashboard[start]
        cell.value = f"{label}\n{value}"
        cell.number_format = number_format
        cell.fill = fill
        cell.font = Font(size=12, bold=True, color="123B5D")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws_dashboard.row_dimensions[5].height = 52

    scenario_image = Image(io.BytesIO(images["dashboard_cenarios.png"]))
    scenario_image.width, scenario_image.height = 880, 475
    ws_dashboard.add_image(scenario_image, "A8")
    ws_dashboard["A33"] = "Recomendação executiva"
    ws_dashboard["A33"].font = Font(size=16, bold=True, color="123B5D")
    ws_dashboard.merge_cells("A34:L37")
    recommendation = (
        "Priorizar a análise do regime híbrido: o relatório do Domínio indica variação de apenas "
        f"{report.fase_2027['diferenca_percentual']:.2%} em 2027 e {report.fase_2033['diferenca_percentual']:.2%} em 2033, "
        f"enquanto aproximadamente {report.percentual_operacoes_creditaveis:.2%} das saídas podem ser sensíveis a crédito. "
        "A decisão final depende da conciliação das entradas, validação por cliente e confirmação das alíquotas aplicáveis."
    )
    ws_dashboard["A34"] = recommendation
    ws_dashboard["A34"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_dashboard["A34"].fill = green_fill
    ws_dashboard["A34"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column in range(1, 13):
        ws_dashboard.column_dimensions[get_column_letter(column)].width = 13

    scenario_headers = list(scenarios.columns)
    ws_scenarios.append(scenario_headers)
    for row in scenarios.itertuples(index=False, name=None):
        ws_scenarios.append(list(row))
    style_header(ws_scenarios, 1, 1, len(scenario_headers))
    for row in range(2, ws_scenarios.max_row + 1):
        for column in (2, 4, 5):
            ws_scenarios.cell(row, column).number_format = money_format
        ws_scenarios.cell(row, 3).number_format = percent_format
    scenario_table_ref = Table(displayName="TabelaCenarios", ref=f"A1:F{ws_scenarios.max_row}")
    scenario_table_ref.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_scenarios.add_table(scenario_table_ref)
    ws_scenarios.freeze_panes = "A2"
    autofit(ws_scenarios, 62)

    assumptions = [
        ("Receita base", report.base_saidas, "Domínio — débitos pelas saídas"),
        ("Base de crédito", report.base_entradas_credito, "Domínio — créditos pelas entradas"),
        ("Alíquota efetiva atual", report.tributos_atuais["Total"] / report.base_saidas, "Carga atual / saídas"),
        ("DAS residual 2033", report.fase_2033["simples_residual"] / report.base_saidas, "Domínio"),
        ("CBS 2033", report.aliquota_cbs_2033, "Domínio"),
        ("IBS 2033", report.aliquota_ibs_2033, "Domínio"),
        ("Receita de contratos em risco", 0.05, "Premissa editável"),
        ("Margem de contribuição", 0.30, "Premissa editável"),
    ]
    ws_assumptions.append(["Premissa", "Valor", "Fonte/observação"])
    for row in assumptions:
        ws_assumptions.append(row)
    style_header(ws_assumptions, 1, 1, 3)
    for row in range(2, ws_assumptions.max_row + 1):
        ws_assumptions.cell(row, 2).fill = yellow_fill if row >= 8 else blue_fill
        if row == 2 or row == 3:
            ws_assumptions.cell(row, 2).number_format = money_format
        else:
            ws_assumptions.cell(row, 2).number_format = percent_format
    autofit(ws_assumptions, 55)

    ws_sensitivity["A1"] = "Sensibilidade — Híbrido 2033 vs. Por Dentro"
    ws_sensitivity["A1"].font = title_font
    ws_sensitivity["A3"] = "Crescimento da receita"
    credit_ratios = [0.50, 0.70, report.base_entradas_credito / report.base_saidas, 1.00]
    for column, ratio in enumerate(credit_ratios, start=2):
        ws_sensitivity.cell(3, column, ratio).number_format = percent_format
    growth_rates = [-0.10, 0.0, 0.10, 0.25]
    for row, growth in enumerate(growth_rates, start=4):
        ws_sensitivity.cell(row, 1, growth).number_format = percent_format
        for column in range(2, 6):
            # Diferença: híbrido (DAS residual + IVA líquido) menos Por Dentro.
            ws_sensitivity.cell(row, column).value = (
                f"=Premissas!$B$2*(1+$A{row})*Premissas!$B$5+"
                f"MAX(Premissas!$B$2*(1+$A{row})*(Premissas!$B$6+Premissas!$B$7)-"
                f"Premissas!$B$2*(1+$A{row})*{get_column_letter(column)}$3*(Premissas!$B$6+Premissas!$B$7),0)-"
                f"Premissas!$B$2*(1+$A{row})*Premissas!$B$4"
            )
            ws_sensitivity.cell(row, column).number_format = money_format
    style_header(ws_sensitivity, 3, 1, 5)
    ws_sensitivity.conditional_formatting.add(
        "B4:E7",
        ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
    )
    ws_sensitivity["A10"] = "Valores negativos favorecem o híbrido; positivos favorecem o Por Dentro, antes dos efeitos comerciais."
    ws_sensitivity["A10"].font = subtitle_font
    autofit(ws_sensitivity, 38)

    ws_monthly.append(["Competência", "Entradas", "Saídas", "Serviços", "Margem bruta de fluxo"])
    for row in monthly.movimentos.itertuples(index=False):
        ws_monthly.append([row[0].to_pydatetime(), row[1], row[2], row[3], row[2] + row[3] - row[1]])
    style_header(ws_monthly, 1, 1, 5)
    for row in range(2, ws_monthly.max_row + 1):
        ws_monthly.cell(row, 1).number_format = "mmm/yyyy"
        for column in range(2, 6):
            ws_monthly.cell(row, column).number_format = money_format
    monthly_table = Table(displayName="TabelaMensal", ref=f"A1:E{ws_monthly.max_row}")
    monthly_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    ws_monthly.add_table(monthly_table)
    line_chart = LineChart()
    line_chart.title = "Entradas x Saídas"
    line_chart.y_axis.title = "R$"
    line_chart.x_axis.title = "Competência"
    line_chart.add_data(Reference(ws_monthly, min_col=2, max_col=3, min_row=1, max_row=ws_monthly.max_row), titles_from_data=True)
    line_chart.set_categories(Reference(ws_monthly, min_col=1, min_row=2, max_row=ws_monthly.max_row))
    line_chart.height, line_chart.width = 9, 18
    ws_monthly.add_chart(line_chart, "G2")
    ws_monthly.freeze_panes = "A2"
    autofit(ws_monthly)

    ws_composition.append(["Componente", "Atual", "Híbrido 2027", "Híbrido 2033"])
    components = ["Simples/DAS residual", "CBS", "IBS", "Total"]
    current_total = report.tributos_atuais["Total"]
    values = [
        [current_total, report.fase_2027["simples_residual"], report.fase_2033["simples_residual"]],
        [0.0, report.fase_2027["cbs"], report.fase_2033["cbs"]],
        [0.0, report.fase_2027["ibs"], report.fase_2033["ibs"]],
        [current_total, report.fase_2027["total"], report.fase_2033["total"]],
    ]
    for component, row_values in zip(components, values):
        ws_composition.append([component, *row_values])
    style_header(ws_composition, 1, 1, 4)
    for row in range(2, ws_composition.max_row + 1):
        for column in range(2, 5):
            ws_composition.cell(row, column).number_format = money_format
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Composição da carga"
    chart.add_data(Reference(ws_composition, min_col=2, max_col=4, min_row=1, max_row=4), titles_from_data=True)
    chart.set_categories(Reference(ws_composition, min_col=1, min_row=2, max_row=4))
    chart.height, chart.width = 9, 17
    ws_composition.add_chart(chart, "F2")
    autofit(ws_composition)

    ws_attention.append(["Prioridade", "Tema", "Constatação", "Ação recomendada"])
    for point in points:
        ws_attention.append([point["Prioridade"], point["Tema"], point["Constatação"], point["Ação"]])
    style_header(ws_attention, 1, 1, 4)
    for row in range(2, ws_attention.max_row + 1):
        priority = ws_attention.cell(row, 1).value
        fill = red_fill if priority == "CRÍTICO" else yellow_fill if priority == "ATENÇÃO" else blue_fill
        ws_attention.cell(row, 1).fill = fill
        ws_attention.cell(row, 1).font = Font(bold=True)
        for column in range(1, 5):
            ws_attention.cell(row, column).alignment = Alignment(wrap_text=True, vertical="top")
        ws_attention.row_dimensions[row].height = 48
    ws_attention.column_dimensions["A"].width = 14
    ws_attention.column_dimensions["B"].width = 28
    ws_attention.column_dimensions["C"].width = 55
    ws_attention.column_dimensions["D"].width = 75
    ws_attention.freeze_panes = "A2"

    raw_sections = [
        ("Saídas por acumulador", report.saidas_por_acumulador),
        ("Entradas por acumulador", report.entradas_por_acumulador),
        ("Clientes por regime", report.clientes_por_regime),
        ("Fornecedores por regime", report.fornecedores_por_regime),
    ]
    current_row = 1
    for title, frame in raw_sections:
        ws_raw.cell(current_row, 1, title).font = Font(size=14, bold=True, color="123B5D")
        current_row += 1
        for column, header in enumerate(frame.columns, start=1):
            ws_raw.cell(current_row, column, header)
        style_header(ws_raw, current_row, 1, max(len(frame.columns), 1))
        for values in frame.itertuples(index=False, name=None):
            current_row += 1
            for column, value in enumerate(values, start=1):
                ws_raw.cell(current_row, column, value)
                if isinstance(value, float):
                    ws_raw.cell(current_row, column).number_format = money_format
        current_row += 3
    autofit(ws_raw, 65)

    if pgdas is not None:
        compatible = cnpjs_are_compatible(report.cnpj, pgdas)
        ws_pgdas.append(["PGDAS-D importado", "Valor"])
        pgdas_rows = [
            ("Empresa", pgdas.empresa),
            ("CNPJ", pgdas.cnpj_estabelecimento or pgdas.cnpj_basico),
            ("Compatível com Domínio", "SIM" if compatible else "NÃO — arquivo de exemplo separado"),
            ("Período", pgdas.periodo),
            ("Anexo", pgdas.anexo),
            ("RPA", pgdas.rpa),
            ("RBT12", pgdas.rbt12),
            ("DAS", pgdas.total_das),
            ("Alíquota efetiva", pgdas.aliquota_efetiva),
            ("Atividade", pgdas.atividade),
        ]
        for row in pgdas_rows:
            ws_pgdas.append(row)
        style_header(ws_pgdas, 1, 1, 2)
        for row in range(2, ws_pgdas.max_row + 1):
            if ws_pgdas.cell(row, 1).value in {"RPA", "RBT12", "DAS"}:
                ws_pgdas.cell(row, 2).number_format = money_format
            if ws_pgdas.cell(row, 1).value == "Alíquota efetiva":
                ws_pgdas.cell(row, 2).number_format = percent_format
        if not compatible:
            ws_pgdas["B4"].fill = red_fill
            ws_pgdas["B4"].font = Font(bold=True, color="9C0006")
        autofit(ws_pgdas, 100)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_excel_dashboard(
    report: DominioSimulationReport,
    monthly: MonthlyReport,
    pgdas: PGDASReport | None = None,
    projection: FutureProjection | None = None,
    intelligent_report: str | None = None,
    reports: Sequence[DominioSimulationReport] | None = None,
    lc214_comparison: pd.DataFrame | None = None,
    lc214_simulation: SimplesLC214Simulation | None = None,
    activity_candidates: pd.DataFrame | None = None,
) -> bytes:
    """Cria um XLSX autocontido usando XlsxWriter, com fórmulas e gráficos."""
    import xlsxwriter

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    workbook.set_properties(
        {
            "title": "Diagnóstico Nascel — Reforma Tributária IBS/CBS",
            "subject": "Simulação gerencial para optante do Simples Nacional",
            "author": NASCEL_NAME,
            "comments": "Premissas devem ser validadas conforme legislação e etapa de transição.",
        }
    )
    workbook.set_calc_mode("auto")

    title = workbook.add_format({"bold": True, "font_size": 22, "font_color": NASCEL_COLORS["navy"], "font_name": "Montserrat"})
    section = workbook.add_format({"bold": True, "font_size": 15, "font_color": NASCEL_COLORS["navy"], "font_name": "Montserrat"})
    subtitle = workbook.add_format({"font_size": 11, "font_color": NASCEL_COLORS["slate"], "font_name": "Montserrat"})
    header = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": NASCEL_COLORS["navy"], "align": "center", "valign": "vcenter", "border": 1, "border_color": NASCEL_COLORS["gold"], "font_name": "Montserrat"}
    )
    text = workbook.add_format({"border": 1, "border_color": "#D7E0E8", "valign": "top"})
    wrap = workbook.add_format({"border": 1, "border_color": "#D7E0E8", "valign": "top", "text_wrap": True})
    money = workbook.add_format({"num_format": 'R$ #,##0.00;[Red]-R$ #,##0.00', "border": 1, "border_color": "#D7E0E8"})
    percent = workbook.add_format({"num_format": "0.00%", "border": 1, "border_color": "#D7E0E8"})
    date_format = workbook.add_format({"num_format": "mmm/yyyy", "border": 1, "border_color": "#D7E0E8"})
    input_percent = workbook.add_format({"num_format": "0.00%", "bg_color": "#FFF4CC", "border": 1, "border_color": "#D7E0E8"})
    note = workbook.add_format({"font_color": NASCEL_COLORS["slate"], "italic": True, "text_wrap": True})
    recommendation_format = workbook.add_format(
        {"bg_color": NASCEL_COLORS["light_gold"], "font_color": NASCEL_COLORS["navy"], "bold": True, "text_wrap": True, "valign": "top", "border": 1, "border_color": NASCEL_COLORS["gold"]}
    )
    critical = workbook.add_format({"bg_color": "#FDE2E0", "font_color": "#9C0006", "bold": True, "border": 1})
    warning = workbook.add_format({"bg_color": "#FFF4CC", "font_color": "#9C6500", "bold": True, "border": 1})
    decision = workbook.add_format({"bg_color": NASCEL_COLORS["light_navy"], "font_color": NASCEL_COLORS["navy"], "bold": True, "border": 1})
    guide_format = workbook.add_format(
        {"bg_color": NASCEL_COLORS["cream"], "font_color": NASCEL_COLORS["navy"], "text_wrap": True, "valign": "vcenter", "border": 1, "border_color": NASCEL_COLORS["gold"]}
    )

    reports = list(reports or [report])
    projection = projection or build_future_projection(
        reports, monthly, lc214_simulation=lc214_simulation
    )
    intelligent_report = intelligent_report or generate_local_intelligent_report(
        projection, reports, monthly, pgdas
    )
    scenarios = scenario_table(report, lc214_simulation)
    points = attention_points(report, monthly, pgdas)
    images = create_dashboard_images(
        report, monthly, pgdas, projection, lc214_simulation
    )
    explanations = dashboard_explanations(report, monthly, pgdas, projection, lc214_simulation)
    diagnostic = (
        build_nascel_diagnostic(report, monthly, pgdas, reports, lc214_simulation)
        if lc214_simulation is not None else None
    )
    decision_matrix = (
        decision_matrix_frame(report, lc214_simulation)
        if lc214_simulation is not None else pd.DataFrame()
    )
    legal_timeline = legal_timeline_frame()
    official_sources = official_sources_frame()

    dashboard = workbook.add_worksheet("Dashboard")
    dashboard.hide_gridlines(2)
    dashboard.set_landscape()
    dashboard.set_paper(9)
    dashboard.fit_to_pages(1, 2)
    dashboard.set_margins(0.3, 0.3, 0.4, 0.4)
    dashboard.set_column("A:L", 13)
    dashboard.merge_range("A1:L2", "Diagnóstico Tributário — Grupo Nascel", title)
    dashboard.merge_range(
        "A3:L3", f"{report.empresa} | CNPJ {report.cnpj} | Competência {report.periodo:%m/%Y}", subtitle
    )

    kpi_formats = []
    for color in (NASCEL_COLORS["light_navy"], NASCEL_COLORS["cream"], NASCEL_COLORS["light_gold"], "#FDE2E0" if report.fase_2027["diferenca"] > 0 else NASCEL_COLORS["cream"]):
        kpi_formats.append(
            workbook.add_format(
                {"bg_color": color, "font_color": NASCEL_COLORS["navy"], "bold": True, "font_size": 12, "align": "center", "valign": "vcenter", "text_wrap": True, "border": 1, "border_color": NASCEL_COLORS["gold"], "font_name": "Montserrat"}
            )
        )
    dashboard.set_row(4, 52)
    dashboard.merge_range("A5:C5", f"Saídas analisadas\nR$ {report.base_saidas:,.2f}", kpi_formats[0])
    dashboard.merge_range("D5:F5", f"Base de entradas\nR$ {report.base_entradas_credito:,.2f}", kpi_formats[1])
    dashboard.merge_range("G5:I5", f"Compras potencialmente creditáveis\n{projection.percentual_entradas_creditaveis:.2%}", kpi_formats[2])
    dashboard.merge_range("J5:L5", f"Diferença 2027\n{report.fase_2027['diferenca_percentual']:.2%}", kpi_formats[3])
    dashboard.set_row(5, 42)
    dashboard.merge_range(
        "A6:L6",
        "Como ler: compare a alíquota e o total a pagar em cada cenário; depois confira o crédito estimado das compras e valide documentos, fornecedores e premissas.",
        guide_format,
    )
    dashboard.insert_image("A8", "dashboard_cenarios.png", {"image_data": io.BytesIO(images["dashboard_cenarios.png"]), "x_scale": 0.62, "y_scale": 0.62})
    dashboard.write("A32", "Recomendação executiva", section)
    recommendation = (
        f"Índice de confiança {diagnostic.score}/100 — {diagnostic.status}. "
        f"{diagnostic.recommendation} {diagnostic.rationale}"
        if diagnostic is not None else
        "Condicionar a decisão à conciliação das entradas, validação por cliente e confirmação das alíquotas aplicáveis."
    )
    dashboard.merge_range("A33:L36", recommendation, recommendation_format)
    dashboard.write("A38", "Projeção dos períodos futuros", section)
    dashboard.insert_image(
        "A40",
        "dashboard_projecao_futura.png",
        {"image_data": io.BytesIO(images["dashboard_projecao_futura.png"]), "x_scale": 0.62, "y_scale": 0.62},
    )

    guide_sheet = workbook.add_worksheet("Como_Ler")
    guide_sheet.freeze_panes(4, 1)
    guide_sheet.hide_gridlines(2)
    guide_sheet.set_column("A:A", 30)
    guide_sheet.set_column("B:B", 55)
    guide_sheet.set_column("C:C", 60)
    guide_sheet.set_column("D:D", 55)
    guide_sheet.set_column("E:E", 65)
    guide_sheet.merge_range("A1:E1", "Como interpretar este dashboard", title)
    guide_sheet.merge_range(
        "A2:E2",
        "O dashboard é uma ferramenta gerencial. Valores importados, cálculos do simulador e premissas são identificados separadamente para facilitar a revisão.",
        guide_format,
    )
    guide_sheet.write_row(3, 0, list(explanations.columns), header)
    for row_index, values in enumerate(explanations.itertuples(index=False, name=None), start=4):
        guide_sheet.set_row(row_index, 72)
        for column_index, value in enumerate(values):
            guide_sheet.write(row_index, column_index, value, wrap)

    if diagnostic is not None:
        diagnostic_sheet = workbook.add_worksheet("Diagnostico_Nascel")
        diagnostic_sheet.hide_gridlines(2)
        diagnostic_sheet.set_column("A:A", 32)
        diagnostic_sheet.set_column("B:C", 12)
        diagnostic_sheet.set_column("D:D", 18)
        diagnostic_sheet.set_column("E:F", 62)
        diagnostic_sheet.merge_range("A1:F1", "Diagnóstico de confiança para a decisão", title)
        diagnostic_sheet.merge_range(
            "A2:F2",
            f"Índice {diagnostic.score}/100 · {diagnostic.status} · {diagnostic.recommendation}",
            recommendation_format,
        )
        diagnostic_sheet.write_row(3, 0, list(diagnostic.checklist.columns), header)
        for row_index, values in enumerate(diagnostic.checklist.itertuples(index=False, name=None), start=4):
            diagnostic_sheet.set_row(row_index, 48)
            for column_index, value in enumerate(values):
                diagnostic_sheet.write(row_index, column_index, value, wrap)

        decision_sheet = workbook.add_worksheet("Decisao_Tributaria")
        decision_sheet.hide_gridlines(2)
        decision_sheet.set_column("A:A", 26)
        decision_sheet.set_column("B:C", 60)
        decision_sheet.set_column("D:E", 24)
        decision_sheet.set_column("F:H", 62)
        decision_sheet.merge_range("A1:H1", "Matriz Nascel — CBS/IBS dentro × fora do DAS", title)
        decision_sheet.merge_range("A2:H2", NASCEL_TAGLINE, guide_format)
        decision_sheet.write_row(3, 0, list(decision_matrix.columns), header)
        for row_index, values in enumerate(decision_matrix.itertuples(index=False, name=None), start=4):
            decision_sheet.set_row(row_index, 92)
            for column_index, value in enumerate(values):
                fmt = money if column_index in {3, 4} else wrap
                decision_sheet.write(row_index, column_index, value, fmt)

    timeline_sheet = workbook.add_worksheet("Cronograma_Legal")
    timeline_sheet.hide_gridlines(2)
    timeline_sheet.set_column("A:A", 16)
    timeline_sheet.set_column("B:D", 62)
    timeline_sheet.set_column("E:E", 54)
    timeline_sheet.merge_range("A1:E1", "Cronograma legal e plano de preparação", title)
    timeline_sheet.merge_range(
        "A2:E2",
        "Base informativa: LC 123/2006 e LC 214/2025, consideradas as alterações posteriores vigentes. Confirmar regulamentação e prazos antes da opção.",
        guide_format,
    )
    timeline_sheet.write_row(3, 0, list(legal_timeline.columns), header)
    for row_index, values in enumerate(legal_timeline.itertuples(index=False, name=None), start=4):
        timeline_sheet.set_row(row_index, 76)
        for column_index, value in enumerate(values):
            timeline_sheet.write(row_index, column_index, value, wrap)

    sources_sheet = workbook.add_worksheet("Fontes_Oficiais")
    sources_sheet.hide_gridlines(2)
    sources_sheet.set_column("A:A", 40)
    sources_sheet.set_column("B:B", 68)
    sources_sheet.set_column("C:C", 92)
    sources_sheet.merge_range("A1:C1", "Fontes oficiais para conferência", title)
    sources_sheet.merge_range(
        "A2:C2",
        "A simulação não congela a legislação. Consulte os textos compilados e a regulamentação vigente antes da decisão.",
        guide_format,
    )
    sources_sheet.write_row(3, 0, list(official_sources.columns), header)
    for row_index, values in enumerate(official_sources.itertuples(index=False, name=None), start=4):
        sources_sheet.set_row(row_index, 46)
        sources_sheet.write(row_index, 0, values[0], wrap)
        sources_sheet.write(row_index, 1, values[1], wrap)
        sources_sheet.write_url(row_index, 2, values[2], text)

    scenario_sheet = workbook.add_worksheet("Cenarios")
    scenario_sheet.freeze_panes(1, 0)
    scenario_sheet.set_column("A:A", 32)
    scenario_sheet.set_column("B:E", 21)
    scenario_sheet.set_column("F:F", 68)
    columns = [
        {"header": "Cenário"}, {"header": "Carga Tributária", "format": money},
        {"header": "Carga Efetiva", "format": percent}, {"header": "Crédito Estimado das Compras", "format": money},
        {"header": "Variação vs. Atual", "format": money}, {"header": "Leitura"},
    ]
    scenario_sheet.add_table(
        0, 0, len(scenarios), len(columns) - 1,
        {"name": "TabelaCenarios", "style": "Table Style Medium 2", "columns": columns, "data": scenarios.values.tolist()},
    )
    chart = workbook.add_chart({"type": "column"})
    chart.add_series(
        {"name": "Carga tributária", "categories": "=Cenarios!$A$2:$A$5", "values": "=Cenarios!$B$2:$B$5", "fill": {"color": NASCEL_COLORS["gold"]}, "border": {"color": NASCEL_COLORS["navy"]}, "data_labels": {"value": True, "num_format": "R$ #,##0"}}
    )
    chart.set_title({"name": "Carga tributária por cenário"})
    chart.set_y_axis({"name": "R$", "num_format": "R$ #,##0"})
    chart.set_legend({"none": True})
    chart.set_size({"width": 820, "height": 420})
    scenario_sheet.insert_chart("A8", chart)

    if lc214_comparison is not None and lc214_simulation is not None:
        lc214_sheet = workbook.add_worksheet("Simples_LC214")
        lc214_sheet.freeze_panes(6, 1)
        lc214_sheet.set_column("A:A", 18)
        lc214_sheet.set_column("B:D", 25)
        lc214_sheet.set_column("E:F", 26)
        lc214_sheet.merge_range(
            "A1:F1", "Simples Nacional 2027/2028 — LC 214/2025", title
        )
        lc214_sheet.write_row(
            2,
            0,
            ["Anexo", "Faixa", "RBT12", "Alíquota nominal", "Taxa teórica da tabela", "Alíquota preservada 2026/2027"],
            header,
        )
        lc214_sheet.write(3, 0, lc214_simulation.annex, text)
        lc214_sheet.write(3, 1, lc214_simulation.bracket, text)
        lc214_sheet.write(3, 2, lc214_simulation.rbt12, money)
        lc214_sheet.write(3, 3, lc214_simulation.nominal_rate, percent)
        lc214_sheet.write(3, 4, lc214_simulation.effective_rate, percent)
        preserved_rate = report.tributos_atuais["Total"] / report.base_saidas if report.base_saidas else 0.0
        lc214_sheet.write(3, 5, preserved_rate, percent)
        lc214_sheet.write_row(5, 0, list(lc214_comparison.columns), header)
        for row_index, values in enumerate(
            lc214_comparison.itertuples(index=False, name=None), start=6
        ):
            for column_index, value in enumerate(values):
                if column_index in {1, 2, 3}:
                    fmt = money
                elif values[0] == "Total":
                    fmt = decision
                else:
                    fmt = wrap if column_index == 4 else text
                lc214_sheet.write(row_index, column_index, value, fmt)
        lc214_sheet.write(
            len(lc214_comparison) + 8,
            0,
            "2027 Por Dentro preserva a carga efetiva de 2026 na mesma base; muda a partilha interna. No Anexo II, o IPI permanece conforme a tabela oficial, sem troca automática de Anexo.",
            note,
        )

    assumptions_sheet = workbook.add_worksheet("Premissas")
    assumptions_sheet.set_column("A:A", 34)
    assumptions_sheet.set_column("B:B", 20)
    assumptions_sheet.set_column("C:C", 58)
    assumptions = [
        ["Receita base", report.base_saidas, "Domínio — débitos pelas saídas"],
        ["Base de crédito", report.base_entradas_credito, "Domínio — créditos pelas entradas"],
        ["Alíquota efetiva atual", report.tributos_atuais["Total"] / report.base_saidas, "Carga atual / saídas"],
        ["DAS residual 2033", report.fase_2033["simples_residual"] / report.base_saidas, "Domínio"],
        ["CBS 2033", report.aliquota_cbs_2033, "Domínio"],
        ["IBS 2033", report.aliquota_ibs_2033, "Domínio"],
        ["Receita de contratos em risco", 0.05, "Premissa editável"],
        ["Margem de contribuição", 0.30, "Premissa editável"],
        ["Modo de crescimento", "Automático pela média" if projection.modo_crescimento == "average" else "Percentual fixo", "Projeção futura"],
        ["Crescimento anual aplicado", projection.crescimento_anual, "Calculado pelo motor de projeção"],
        ["Crescimento mensal médio", projection.crescimento_mensal_medio, f"Últimos {projection.meses_calculo_crescimento} períodos"],
    ]
    assumptions_sheet.write_row(0, 0, ["Premissa", "Valor", "Fonte/observação"], header)
    for row_index, row in enumerate(assumptions, start=1):
        assumptions_sheet.write(row_index, 0, row[0], text)
        if row[0] in {"Receita base", "Base de crédito"}:
            fmt = money
        elif row[0] in {"Receita de contratos em risco", "Margem de contribuição"}:
            fmt = input_percent
        elif isinstance(row[1], (int, float)):
            fmt = percent
        else:
            fmt = text
        assumptions_sheet.write(row_index, 1, row[1], fmt)
        assumptions_sheet.write(row_index, 2, row[2], wrap)
    assumptions_sheet.data_validation("B8:B9", {"validate": "decimal", "criteria": "between", "minimum": 0, "maximum": 1, "input_title": "Premissa editável", "input_message": "Informe um percentual entre 0% e 100%."})

    sensitivity = workbook.add_worksheet("Sensibilidade")
    sensitivity.hide_gridlines(2)
    sensitivity.freeze_panes(9, 0)
    sensitivity.set_column("A:A", 31)
    sensitivity.set_column("B:B", 20)
    sensitivity.set_column("C:I", 22)
    sensitivity.set_column("J:J", 31)
    sensitivity.set_row(1, 28)
    sensitivity.set_row(2, 28)
    sensitivity.set_row(4, 28)
    sensitivity.set_row(5, 28)
    sensitivity.set_row(6, 32)
    sensitivity.set_row(8, 42)
    for sensitivity_row in range(9, 13):
        sensitivity.set_row(sensitivity_row, 30)
    for sensitivity_row in range(14, 17):
        sensitivity.set_row(sensitivity_row, 24)
    sensitivity.merge_range(
        "A1:J1", "Sensibilidade das compras — Híbrido 2033 vs. DAS Normal", title
    )
    sensitivity.merge_range(
        "A2:J3",
        "FINALIDADE: verificar se a decisão muda quando a base de compras apta a gerar crédito de IBS/CBS fica menor ou maior. "
        "A receita anual e a alíquota do DAS Normal permanecem iguais à projeção de 2027; varia somente a base creditável das compras.",
        guide_format,
    )

    annual_2033 = projection.resumo_anual.iloc[1]
    annual_revenue = float(annual_2033["Receita Projetada"])
    normal_total = float(annual_2033["DAS Normal · Valor"])
    current_credit_ratio = (
        float(annual_2033["Base Creditável das Compras"]) / annual_revenue
        if annual_revenue
        else 0.0
    )
    report_output_base = sum(item.base_saidas for item in reports)
    report_input_base = sum(item.base_entradas_credito for item in reports)

    def output_rate(attribute: str) -> float:
        return (
            sum(item.base_saidas * getattr(item, attribute) for item in reports)
            / report_output_base
            if report_output_base
            else 0.0
        )

    def input_rate(attribute: str) -> float:
        return (
            sum(item.base_entradas_credito * getattr(item, attribute) for item in reports)
            / report_input_base
            if report_input_base
            else 0.0
        )

    residual_rate = (
        sum(item.fase_2033["simples_residual"] for item in reports)
        / report_output_base
        if report_output_base
        else 0.0
    )
    output_cbs_rate = output_rate("aliquota_cbs_2033")
    output_ibs_rate = output_rate("aliquota_ibs_2033")
    input_cbs_rate = input_rate("aliquota_credito_cbs_2033")
    input_ibs_rate = input_rate("aliquota_credito_ibs_2033")

    def sensitivity_values(ratio: float) -> dict[str, float]:
        credit_base = annual_revenue * ratio
        purchase_credit = credit_base * (input_cbs_rate + input_ibs_rate)
        residual = annual_revenue * residual_rate
        cbs = max(annual_revenue * output_cbs_rate - credit_base * input_cbs_rate, 0.0)
        ibs = max(annual_revenue * output_ibs_rate - credit_base * input_ibs_rate, 0.0)
        hybrid = residual + cbs + ibs
        return {
            "base": credit_base,
            "credit": purchase_credit,
            "residual": residual,
            "cbs": cbs,
            "ibs": ibs,
            "hybrid": hybrid,
            "difference": hybrid - normal_total,
        }

    low, high = 0.0, 2.0
    if sensitivity_values(high)["difference"] <= 0:
        for _ in range(80):
            middle = (low + high) / 2
            if sensitivity_values(middle)["difference"] > 0:
                low = middle
            else:
                high = middle
        break_even_ratio = high
    else:
        break_even_ratio = None

    current_values = sensitivity_values(current_credit_ratio)
    current_reading = (
        f"Híbrido menor em R$ {abs(current_values['difference']):,.2f}"
        if current_values["difference"] < 0
        else f"DAS Normal menor em R$ {current_values['difference']:,.2f}"
        if current_values["difference"] > 0
        else "Empate entre os regimes"
    )
    sensitivity.merge_range("A5:C6", f"Base creditável atual\n{current_credit_ratio:.2%} da receita", decision)
    sensitivity.merge_range(
        "D5:F6",
        f"Ponto de equilíbrio\n{break_even_ratio:.2%} da receita"
        if break_even_ratio is not None
        else "Ponto de equilíbrio\nAcima de 200% da receita",
        warning,
    )
    sensitivity.merge_range("G5:J6", f"Leitura da base atual\n{current_reading}", recommendation_format)
    sensitivity.merge_range(
        "A7:J7",
        "COMO LER: diferença negativa = Híbrido paga menos; diferença positiva = DAS Normal paga menos. "
        "A linha 'Base atual da empresa' reproduz a estrutura de compras importada.",
        note,
    )

    sensitivity_headers = [
        "Cenário",
        "Base Creditável / Receita",
        "Base Creditável (R$)",
        "Crédito das Compras (R$)",
        "DAS Residual (R$)",
        "CBS Líquida (R$)",
        "IBS Líquido (R$)",
        "Total Híbrido (R$)",
        "Diferença vs. DAS Normal",
        "Interpretação",
    ]
    sensitivity.write_row(8, 0, sensitivity_headers, header)
    ratio_scenarios = [
        ("Compras creditáveis baixas", 0.50),
        ("Compras creditáveis médias", 0.70),
        ("Base atual da empresa", current_credit_ratio),
        ("Compras creditáveis altas", 1.00),
    ]
    for row_index, (label, ratio) in enumerate(ratio_scenarios, start=9):
        values = sensitivity_values(ratio)
        interpretation = (
            "Híbrido paga menos"
            if values["difference"] < 0
            else "DAS Normal paga menos"
            if values["difference"] > 0
            else "Empate"
        )
        sensitivity.write(row_index, 0, label, decision if label == "Base atual da empresa" else text)
        sensitivity.write(row_index, 1, ratio, percent)
        for column, key in enumerate(
            ("base", "credit", "residual", "cbs", "ibs", "hybrid", "difference"),
            start=2,
        ):
            sensitivity.write(row_index, column, values[key], money)
        sensitivity.write(row_index, 9, interpretation, wrap)
    sensitivity.conditional_format(
        "I10:I13",
        {"type": "cell", "criteria": "<", "value": 0, "format": workbook.add_format({"bg_color": "#C6EFCE", "font_color": "#006100", "num_format": 'R$ #,##0.00;[Red]-R$ #,##0.00'})},
    )
    sensitivity.conditional_format(
        "I10:I13",
        {"type": "cell", "criteria": ">", "value": 0, "format": workbook.add_format({"bg_color": "#FFC7CE", "font_color": "#9C0006", "num_format": 'R$ #,##0.00;[Red]-R$ #,##0.00'})},
    )
    sensitivity.merge_range(
        "A15:J17",
        "LIMITAÇÃO: esta aba é uma análise gerencial. O crédito efetivo depende de documento fiscal idôneo, "
        "extinção do débito, vedações, reduções e demais regras aplicáveis. CBS e IBS são calculados separadamente.",
        note,
    )

    projection_sheet = workbook.add_worksheet("Projecao_Futura")
    projection_sheet.freeze_panes(1, 0)
    projection_sheet.set_column("A:A", 20)
    projection_sheet.set_column("B:M", 24)
    projection_columns = list(projection.resumo_anual.columns)
    projection_sheet.write_row(0, 0, projection_columns, header)
    for row_index, values in enumerate(
        projection.resumo_anual.itertuples(index=False, name=None), start=1
    ):
        for column, value in enumerate(values):
            column_name = projection_columns[column]
            cell_format = (
                text
                if column_name == "Período"
                else percent
                if "Alíquota" in column_name
                else money
            )
            projection_sheet.write(
                row_index,
                column,
                value,
                cell_format,
            )
    projection_sheet.add_table(
        0,
        0,
        len(projection.resumo_anual),
        len(projection_columns) - 1,
        {
            "name": "TabelaProjecaoFutura",
            "style": "Table Style Medium 9",
            "columns": [{"header": column} for column in projection_columns],
        },
    )
    projection_chart = workbook.add_chart({"type": "column"})
    for column_name, color in (
        ("DAS Normal · Valor", NASCEL_COLORS["navy"]),
        ("Híbrido · Total a Pagar", NASCEL_COLORS["gold"]),
    ):
        column = projection_columns.index(column_name)
        projection_chart.add_series(
            {
                "name": ["Projecao_Futura", 0, column],
                "categories": ["Projecao_Futura", 1, 0, len(projection.resumo_anual), 0],
                "values": ["Projecao_Futura", 1, column, len(projection.resumo_anual), column],
                "fill": {"color": color},
            }
        )
    projection_chart.set_title({"name": "Total anual: DAS Normal x Híbrido"})
    projection_chart.set_y_axis({"num_format": "R$ #,##0"})
    projection_chart.set_size({"width": 900, "height": 420})
    projection_sheet.insert_chart("A6", projection_chart)

    history_sheet = workbook.add_worksheet("Historico_Simulacoes")
    history_sheet.freeze_panes(1, 0)
    history_sheet.set_column("A:A", 16)
    history_sheet.set_column("B:G", 22)
    history_columns = list(projection.historico_simulacoes.columns)
    history_sheet.write_row(0, 0, history_columns, header)
    for row_index, values in enumerate(
        projection.historico_simulacoes.itertuples(index=False, name=None), start=1
    ):
        for column, value in enumerate(values):
            fmt = date_format if column == 0 else percent if history_columns[column] == "Operações Creditáveis" else money
            history_sheet.write(row_index, column, value.to_pydatetime() if column == 0 else value, fmt)

    ai_sheet = workbook.add_worksheet("Relatorio_Inteligente")
    ai_sheet.hide_gridlines(2)
    ai_sheet.set_column("A:H", 16)
    ai_sheet.merge_range("A1:H2", "Relatório Consultivo Nascel", title)
    ai_sheet.merge_range("A4:H45", intelligent_report, workbook.add_format({"text_wrap": True, "valign": "top", "font_size": 11, "border": 1, "border_color": "#D7E0E8"}))

    monthly_sheet = workbook.add_worksheet("Evolucao_Mensal")
    monthly_sheet.freeze_panes(1, 0)
    monthly_sheet.set_column("A:A", 16)
    monthly_sheet.set_column("B:E", 18)
    monthly_sheet.add_table(
        0, 0, len(monthly.movimentos), 4,
        {
            "name": "TabelaMensal", "style": "Table Style Medium 4",
            "columns": [{"header": "Competência", "format": date_format}, {"header": "Entradas", "format": money}, {"header": "Saídas", "format": money}, {"header": "Serviços", "format": money}, {"header": "Margem bruta de fluxo", "format": money}],
            "data": [[row[0].to_pydatetime(), row[1], row[2], row[3], row[2] + row[3] - row[1]] for row in monthly.movimentos.itertuples(index=False)],
        },
    )
    line = workbook.add_chart({"type": "line"})
    for column, color in ((2, NASCEL_COLORS["gold"]), (3, NASCEL_COLORS["navy"])):
        line.add_series({"name": ["Evolucao_Mensal", 0, column - 1], "categories": ["Evolucao_Mensal", 1, 0, len(monthly.movimentos), 0], "values": ["Evolucao_Mensal", 1, column - 1, len(monthly.movimentos), column - 1], "line": {"color": color, "width": 2.25}, "marker": {"type": "circle", "size": 5}})
    line.set_title({"name": "Evolução mensal de entradas e saídas"})
    line.set_y_axis({"num_format": "R$ #,##0"})
    line.set_size({"width": 900, "height": 420})
    monthly_sheet.insert_chart("G2", line)

    composition = workbook.add_worksheet("Composicao")
    composition.set_column("A:A", 28)
    composition.set_column("B:D", 20)
    composition.write_row(0, 0, ["Componente", "Atual", "Híbrido 2027", "Híbrido 2033"], header)
    current_total = report.tributos_atuais["Total"]
    composition_rows = [
        ["Simples/DAS residual", current_total, report.fase_2027["simples_residual"], report.fase_2033["simples_residual"]],
        ["CBS", 0, report.fase_2027["cbs"], report.fase_2033["cbs"]],
        ["IBS", 0, report.fase_2027["ibs"], report.fase_2033["ibs"]],
        ["Total", current_total, report.fase_2027["total"], report.fase_2033["total"]],
    ]
    for row_index, row in enumerate(composition_rows, start=1):
        composition.write(row_index, 0, row[0], text)
        composition.write_row(row_index, 1, row[1:], money)
    comp_chart = workbook.add_chart({"type": "column", "subtype": "stacked"})
    for column in range(1, 4):
        comp_chart.add_series({"name": ["Composicao", 0, column], "categories": "=Composicao!$A$2:$A$4", "values": ["Composicao", 1, column, 3, column]})
    comp_chart.set_title({"name": "Composição por cenário"})
    comp_chart.set_y_axis({"num_format": "R$ #,##0"})
    comp_chart.set_size({"width": 780, "height": 420})
    composition.insert_chart("F2", comp_chart)

    attention = workbook.add_worksheet("Pontos_Atencao")
    attention.freeze_panes(1, 0)
    attention.set_column("A:A", 14)
    attention.set_column("B:B", 28)
    attention.set_column("C:D", 58)
    attention.set_column("E:E", 55)
    attention.set_column("F:F", 75)
    attention.write_row(
        0, 0,
        ["Prioridade", "Tema", "O que significa", "Dados considerados", "Constatação", "Ação recomendada"],
        header,
    )
    priority_formats = {"CRÍTICO": critical, "ATENÇÃO": warning, "DECISÃO": decision}
    for row_index, point in enumerate(points, start=1):
        attention.set_row(row_index, 82)
        attention.write(row_index, 0, point["Prioridade"], priority_formats.get(point["Prioridade"], text))
        attention.write(row_index, 1, point["Tema"], wrap)
        attention.write(row_index, 2, point["O que significa"], wrap)
        attention.write(row_index, 3, point["Dados considerados"], wrap)
        attention.write(row_index, 4, point["Constatação"], wrap)
        attention.write(row_index, 5, point["Ação"], wrap)

    raw = workbook.add_worksheet("Dados_Dominio")
    raw.set_column("A:A", 62)
    raw.set_column("B:C", 20)
    current_row = 0
    for section_name, frame in (
        ("Saídas por acumulador", report.saidas_por_acumulador),
        ("Entradas por acumulador", report.entradas_por_acumulador),
        ("Clientes por regime", report.clientes_por_regime),
        ("Fornecedores por regime", report.fornecedores_por_regime),
    ):
        raw.write(current_row, 0, section_name, section)
        current_row += 1
        raw.write_row(current_row, 0, list(frame.columns), header)
        current_row += 1
        for values in frame.itertuples(index=False, name=None):
            for column, value in enumerate(values):
                fmt = percent if frame.columns[column] == "Percentual" else money if isinstance(value, float) else text
                raw.write(current_row, column, value, fmt)
            current_row += 1
        current_row += 2

    if pgdas is not None:
        pgdas_sheet = workbook.add_worksheet("PGDAS_Exemplo")
        pgdas_sheet.set_column("A:A", 30)
        pgdas_sheet.set_column("B:B", 95)
        compatible = cnpjs_are_compatible(report.cnpj, pgdas)
        pgdas_sheet.write_row(0, 0, ["PGDAS-D importado", "Valor"], header)
        pgdas_rows = [
            ["Empresa", pgdas.empresa], ["CNPJ", pgdas.cnpj_estabelecimento or pgdas.cnpj_basico],
            ["Compatível com Domínio", "SIM" if compatible else "NÃO — arquivo de exemplo separado"],
            ["Período", pgdas.periodo], ["Anexo", pgdas.anexo], ["RPA", pgdas.rpa], ["RBT12", pgdas.rbt12],
            ["DAS", pgdas.total_das], ["Alíquota efetiva", pgdas.aliquota_efetiva], ["Atividade", pgdas.atividade],
        ]
        for row_index, row in enumerate(pgdas_rows, start=1):
            pgdas_sheet.write(row_index, 0, row[0], text)
            if row[0] in {"RPA", "RBT12", "DAS"}:
                fmt = money
            elif row[0] == "Alíquota efetiva":
                fmt = percent
            elif row[0] == "Compatível com Domínio" and not compatible:
                fmt = critical
            else:
                fmt = wrap
            pgdas_sheet.write(row_index, 1, row[1], fmt)

    if activity_candidates is not None and not activity_candidates.empty:
        activity_sheet = workbook.add_worksheet("Atividades_CNPJ")
        activity_sheet.hide_gridlines(2)
        activity_sheet.freeze_panes(5, 0)
        activity_sheet.merge_range("A1:J2", "Candidatos de atividade e tratamento IBS/CBS", title)
        activity_sheet.merge_range(
            "A3:J3",
            "Resultado indicativo: o CNAE auxilia a localizar possibilidades, mas não determina isoladamente a NBS, a cClassTrib ou a redução aplicável.",
            note,
        )
        candidate_columns = list(activity_candidates.columns)
        activity_sheet.write_row(4, 0, candidate_columns, header)
        for row_index, values in enumerate(
            activity_candidates.itertuples(index=False, name=None), start=5
        ):
            for column_index, value in enumerate(values):
                if pd.isna(value):
                    value = ""
                elif hasattr(value, "item"):
                    value = value.item()
                activity_sheet.write(row_index, column_index, value, wrap)
        activity_sheet.autofilter(4, 0, 4 + len(activity_candidates), len(candidate_columns) - 1)
        activity_sheet.set_column(0, 0, 15)
        activity_sheet.set_column(1, 1, 45)
        activity_sheet.set_column(2, 5, 18)
        activity_sheet.set_column(6, 6, 55)
        activity_sheet.set_column(7, 8, 18)
        activity_sheet.set_column(9, 9, 38)

    workbook.close()
    return output.getvalue()


def build_transactional_template() -> bytes:
    """Modelo XLSX com as três abas aceitas pelo modo transacional."""
    import xlsxwriter

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    title = workbook.add_format({"bold": True, "font_size": 18, "font_color": NASCEL_COLORS["navy"], "font_name": "Montserrat"})
    header = workbook.add_format({"bold": True, "font_color": "white", "bg_color": NASCEL_COLORS["navy"], "border": 1, "border_color": NASCEL_COLORS["gold"], "align": "center", "font_name": "Montserrat"})
    input_format = workbook.add_format({"bg_color": "#FFF4CC", "border": 1})
    date_format = workbook.add_format({"num_format": "dd/mm/yyyy", "bg_color": "#FFF4CC", "border": 1})
    money_format = workbook.add_format({"num_format": 'R$ #,##0.00', "bg_color": "#FFF4CC", "border": 1})
    percent_format = workbook.add_format({"num_format": "0.00%", "bg_color": "#FFF4CC", "border": 1})
    note = workbook.add_format({"text_wrap": True, "valign": "top", "font_color": NASCEL_COLORS["slate"]})

    instructions = workbook.add_worksheet("Instrucoes")
    instructions.set_column("A:A", 28)
    instructions.set_column("B:B", 100)
    instructions.merge_range("A1:B2", "Modelo de Importação — Grupo Nascel", title)
    instruction_rows = [
        ("Faturamento", "Uma linha por documento/operação. Preserve zeros à esquerda no CPF/CNPJ."),
        ("Entradas", "Informe a base potencial de crédito já conciliada; o portal ainda sinalizará a necessidade de validação fiscal."),
        ("Parametros_SN", "Use apenas uma linha válida. A alíquota pode ser informada como 8,50% ou 0,085."),
        ("Datas", "Utilize datas reais do Excel ou o formato dd/mm/aaaa."),
        ("Valores", "Utilize números, sem fórmulas externas. Valores negativos serão tratados como devoluções/ajustes."),
        ("Cabeçalhos", "Não altere os nomes das colunas. Não inclua títulos antes da primeira linha."),
    ]
    instructions.write_row(3, 0, ["Aba", "Orientação"], header)
    for row_index, row in enumerate(instruction_rows, start=4):
        instructions.write(row_index, 0, row[0], input_format)
        instructions.write(row_index, 1, row[1], note)

    sales = workbook.add_worksheet("Faturamento")
    sales.set_column("A:A", 15)
    sales.set_column("B:B", 12)
    sales.set_column("C:C", 24)
    sales.set_column("D:D", 18)
    sales.write_row(0, 0, ["Data", "CFOP", "CNPJ_CPF_Cliente", "Valor_Total"], header)
    sales.write_datetime(1, 0, pd.Timestamp("2026-05-02").to_pydatetime(), date_format)
    sales.write(1, 1, "5102", input_format)
    sales.write(1, 2, "12.345.678/0001-90", input_format)
    sales.write_number(1, 3, 10000.00, money_format)
    sales.autofilter(0, 0, 1000, 3)
    sales.freeze_panes(1, 0)

    purchases = workbook.add_worksheet("Entradas")
    purchases.set_column("A:A", 15)
    purchases.set_column("B:B", 12)
    purchases.set_column("C:D", 22)
    purchases.write_row(0, 0, ["Data", "CFOP", "Valor_Total_Nota", "Valor_Base_Credito"], header)
    purchases.write_datetime(1, 0, pd.Timestamp("2026-05-03").to_pydatetime(), date_format)
    purchases.write(1, 1, "1102", input_format)
    purchases.write_number(1, 2, 7500.00, money_format)
    purchases.write_number(1, 3, 7000.00, money_format)
    purchases.autofilter(0, 0, 1000, 3)
    purchases.freeze_panes(1, 0)

    parameters = workbook.add_worksheet("Parametros_SN")
    parameters.set_column("A:A", 20)
    parameters.set_column("B:B", 14)
    parameters.set_column("C:C", 26)
    parameters.write_row(0, 0, ["RBT12", "Anexo", "Aliquota_Efetiva_Atual"], header)
    parameters.write_number(1, 0, 1200000.00, money_format)
    parameters.write(1, 1, "I", input_format)
    parameters.write_number(1, 2, 0.085, percent_format)
    parameters.data_validation(1, 1, 100, 1, {"validate": "list", "source": ["I", "II", "III", "IV", "V"]})
    parameters.data_validation(1, 2, 100, 2, {"validate": "decimal", "criteria": "between", "minimum": 0, "maximum": 1})
    parameters.freeze_panes(1, 0)

    workbook.close()
    return output.getvalue()
